"""
Drives an insight-outline (see outline.py) into a finished xlsx workbook via
xlsx_insight_lib's styling primitives, adding only the column-anchor math
needed to place an n-way "|"-split image row side by side.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter

from src.report.outline import Section
from src.report.xlsx_insight_lib import (
    COL_PAD_PX,
    DEFAULT_COL_WIDTH,
    MDW,
    bullet,
    insert_image,
    new_workbook,
    save,
)

BASE_WIDTH_PX = 1100  # matches the original skill's widest "chart-below" pattern
IMAGE_ROW_GAP_PX = 16  # horizontal gap between split-row images
START_COL = "C"  # anchor column for a finding's first image slot


def _pixel_width(ws, col: str) -> float:
    """Column's rendered pixel width: ws.column_dimensions override, else
    DEFAULT_COL_WIDTH, via the standard width_units*MDW+COL_PAD_PX formula."""
    dim = ws.column_dimensions[col] if col in ws.column_dimensions else None
    width_units = dim.width if dim is not None and dim.width else DEFAULT_COL_WIDTH
    return width_units * MDW + COL_PAD_PX


def _column_after(ws, start_col: str, offset_px: float) -> str:
    """Walk columns rightward from start_col accumulating pixel widths; return the
    first column letter whose left edge is >= offset_px past start_col."""
    col_idx = column_index_from_string(start_col)
    left_edge = 0.0
    while left_edge < offset_px:
        left_edge += _pixel_width(ws, get_column_letter(col_idx))
        col_idx += 1
    return get_column_letter(col_idx)


def _split_row_columns(
    ws,
    n: int,
    *,
    start_col: str = START_COL,
    base_width_px: int = BASE_WIDTH_PX,
    gap_px: int = IMAGE_ROW_GAP_PX,
) -> list[str]:
    """Concrete 1/2/3(+)-way column anchors: slot 0 = start_col; slot i>0 is
    offset i*(per-slot width + gap) pixels to the right of start_col."""
    slot_width = base_width_px / n
    return [
        start_col if i == 0 else _column_after(ws, start_col, i * (slot_width + gap_px))
        for i in range(n)
    ]


def _place_image_row(
    ws,
    row: int,
    cells: list[Path | None],
    *,
    headline: str,
    base_width_px: int = BASE_WIDTH_PX,
    start_col: str = START_COL,
) -> int:
    """Place one image row: n = len(cells), each slot's target_width_px =
    base_width_px / n. None cells render an italic '[Chart to insert: ...]'
    placeholder instead of an image. Returns the row past the tallest slot."""
    n = len(cells)
    columns = _split_row_columns(ws, n, start_col=start_col, base_width_px=base_width_px)
    target_width_px = int(base_width_px / n)

    next_rows = []
    for i, (col, cell) in enumerate(zip(columns, cells)):
        if cell is None:
            placeholder = f"[Chart to insert: {headline} — slot {i + 1}/{n}]"
            next_rows.append(bullet(ws, row, col, placeholder, italic=True))
        else:
            next_rows.append(insert_image(ws, row, col, cell, target_width_px=target_width_px))
    return max(next_rows)


def build_insight_xlsx(sections: list[Section], out_path: str | Path, *, base_width_px: int = BASE_WIDTH_PX) -> Path:
    """Lay out each finding as a bold headline bullet followed by its image
    row(s), then save. No section_header() bands — the outline-driven flow is
    a flat list of findings, not the original skill's fixed Goal/Info/Results
    skeleton (that skeleton was never fixed to begin with, per its own spec)."""
    wb, ws = new_workbook()
    row = 2
    for section in sections:
        row = bullet(ws, row, "C", section.headline, bold=True)
        for cells in section.rows:
            row = _place_image_row(ws, row, cells, headline=section.headline, base_width_px=base_width_px)
            row += 2  # style-spec: blank rows after an image block
        row += 1  # blank row before next finding's headline
    save(wb, out_path)
    return Path(out_path)
