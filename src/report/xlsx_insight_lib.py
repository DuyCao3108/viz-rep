"""
Styling primitives for building an insight-result workbook, ported from
libs/.claude/skills/insight-draft/references/xlsx_insight_lib.py (see
viz-rep's own .claude/skills/insight-draft/references/style-spec.md for the
full spec these functions encode).

Unlike the original (which lived only as a scratchpad-imported reference),
this is a real, imported, tested module — src/report/build.py drives it
directly instead of a one-off per-session driver script.

Typical usage:

    from src.report.xlsx_insight_lib import new_workbook, bullet, insert_image, save

    wb, ws = new_workbook()
    row = bullet(ws, 2, "C", "1-How many card users we acquired from rivals every month?")
    row = insert_image(ws, row, "C", "/path/to/chart.png", target_width_px=900)

    save(wb, "/path/to/InsightResult_Topic_Analysis.xlsx")
"""

from __future__ import annotations

import os
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import column_index_from_string

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None


SHEET_NAME = "Analysis"
LAST_BAND_COL = "AE"  # section-header bands always fill out to this column

FONT_NAME = "Calibri"
FONT_SIZE = 11
FOOTNOTE_SIZE = 10

FONT_NORMAL = Font(name=FONT_NAME, size=FONT_SIZE)
FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
FONT_ITALIC_UNDERLINE = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, underline="single")
FONT_ITALIC = Font(name=FONT_NAME, size=FONT_SIZE, italic=True)
FONT_FOOTNOTE = Font(name=FONT_NAME, size=FOOTNOTE_SIZE, italic=True)

HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_THIN = Side(style="thin")
BORDER_BAND_LEFT = Border(top=_THIN, bottom=_THIN, left=_THIN)
BORDER_BAND_MID = Border(top=_THIN, bottom=_THIN)
BORDER_BAND_RIGHT = Border(top=_THIN, bottom=_THIN, right=_THIN)
BORDER_TABLE_HEADER_FIRST = Border(bottom=_THIN)
BORDER_TABLE_HEADER = Border(bottom=_THIN, left=_THIN)
BORDER_TABLE_ROW = Border(left=_THIN)

IMAGE_BORDER_COLOR = "D9D9D9"  # light grey frame around every pasted chart PNG

COLUMN_WIDTHS = {"A": 3.1, "B": 4.7, "C": 3.9, "D": 4.0, "E": 3.4}

# Excel default row height (14.4pt) rendered at 96dpi ≈ 19.2px; used to estimate
# how many rows a pasted image will visually span.
PX_PER_ROW = 19.2

# Standard Excel column-width-units -> pixels formula for the default font
# (Calibri 11, max digit width 7px): pixels = round(width_units * 7 + 5).
# Used by build.py's n-way-split column-anchor math to know how many columns
# a given pixel offset actually spans.
MDW = 7
COL_PAD_PX = 5
DEFAULT_COL_WIDTH = 8.43


def new_workbook() -> tuple[Workbook, "Worksheet"]:
    """Create the workbook shell: single 'Analysis' sheet, gridlines off, indent columns set."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    return wb, ws


def section_header(ws, row: int, title: str, last_col: str = LAST_BAND_COL) -> int:
    """Write a full-width section band (e.g. '1 Goal is to answer') at `row`. Returns next row."""
    last_idx = column_index_from_string(last_col)
    for col_idx in range(column_index_from_string("B"), last_idx + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        if col_idx == column_index_from_string("B"):
            cell.border = BORDER_BAND_LEFT
            cell.value = title
            cell.font = FONT_BOLD
        elif col_idx == last_idx:
            cell.border = BORDER_BAND_RIGHT
        else:
            cell.border = BORDER_BAND_MID
    return row + 1


def subheader(ws, row: int, text: str, col: str = "C") -> int:
    """Write an italic+underline topic label (e.g. 'Data sources'). Returns next row."""
    cell = ws.cell(row=row, column=column_index_from_string(col), value=text)
    cell.font = FONT_ITALIC_UNDERLINE
    return row + 1


def bullet(ws, row: int, col: str, text: str, bold: bool = False, italic: bool = False) -> int:
    """Write a plain (or bold/italic) body line at the given indent column. Returns next row."""
    cell = ws.cell(row=row, column=column_index_from_string(col), value=text)
    if bold:
        cell.font = FONT_BOLD
    elif italic:
        cell.font = FONT_ITALIC
    else:
        cell.font = FONT_NORMAL
    return row + 1


def footnote(ws, row: int, col: str, text: str, marker: str = "(*) ") -> int:
    """Write a small italic footnote, prefixed with `marker`. Returns next row."""
    cell = ws.cell(row=row, column=column_index_from_string(col), value=marker + text)
    cell.font = FONT_FOOTNOTE
    return row + 1


def hypothesis_table(
    ws,
    row: int,
    rows: list[dict],
    num_col: str = "D",
    hyp_col: str = "E",
    cond_col: str = "M",
) -> int:
    """
    Write a hypothesis mini-table: '#' / 'Hypothesis' / 'Hypothesis is true if' header,
    then one row per dict with keys num, hypothesis, condition. Returns next row.
    """
    headers = [(num_col, "#", "right"), (hyp_col, "Hypothesis", None), (cond_col, "Hypothesis is true if", None)]
    for col, label, align in headers:
        cell = ws.cell(row=row, column=column_index_from_string(col), value=label)
        cell.font = FONT_ITALIC
        cell.border = BORDER_TABLE_HEADER_FIRST if col == num_col else BORDER_TABLE_HEADER
        if align:
            cell.alignment = Alignment(horizontal=align)
    row += 1
    for r in rows:
        num_cell = ws.cell(row=row, column=column_index_from_string(num_col), value=r["num"])
        num_cell.font = FONT_NORMAL
        hyp_cell = ws.cell(row=row, column=column_index_from_string(hyp_col), value=r["hypothesis"])
        hyp_cell.font = FONT_NORMAL
        hyp_cell.border = BORDER_TABLE_ROW
        cond_cell = ws.cell(row=row, column=column_index_from_string(cond_col), value=r["condition"])
        cond_cell.font = FONT_NORMAL
        cond_cell.border = BORDER_TABLE_ROW
        row += 1
    return row


def estimate_image_rows(image_path: str | Path, target_width_px: int) -> tuple[int, int]:
    """
    Return (scaled_height_px, row_span) for an image scaled to target_width_px,
    preserving aspect ratio. Requires Pillow; raises if unavailable.
    """
    if PILImage is None:
        raise RuntimeError("Pillow is required to estimate image row span (pip install pillow)")
    with PILImage.open(image_path) as im:
        w, h = im.size
    scaled_height = int(h * (target_width_px / w))
    row_span = max(1, round(scaled_height / PX_PER_ROW))
    return scaled_height, row_span


def insert_image(
    ws,
    row: int,
    col: str,
    image_path: str | Path,
    target_width_px: int = 900,
) -> int:
    """
    Paste a chart PNG at (row, col), scaled to target_width_px wide (aspect preserved).
    Returns the row immediately after the image's estimated bottom edge, so the caller
    can add the "2-3 blank row" gap before the next element per the style spec.

    Every pasted picture gets a native Picture Border (the same "Picture Format >
    Picture Border" outline you'd set by hand in Excel) applied as a post-processing
    pass in `save()` — openpyxl's high-level Image API has no property for this, so
    it's done by patching the drawing XML after `wb.save()`. Nothing to do here.
    """
    img = XLImage(str(image_path))
    scaled_height, row_span = estimate_image_rows(image_path, target_width_px)
    img.width = target_width_px
    img.height = scaled_height
    anchor = f"{col}{row}"
    ws.add_image(img, anchor)
    return row + row_span


def _add_picture_borders(path: str, color: str = IMAGE_BORDER_COLOR, width_emu: int = 9525) -> None:
    """
    Patch every pasted picture's drawing XML in the saved xlsx to add a native
    Picture Border (a:ln with a solid fill), matching what Excel's own
    Picture Format > Picture Border color picker would set. openpyxl's Image/
    PictureFrame writer always emits `<spPr><a:prstGeom prst="rect"/></spPr>`
    with no outline and no hook to customize it, so this rewrites the zip entry
    for each `xl/drawings/drawingN.xml` part after the workbook is saved.
    """
    ln_xml = f'<a:ln w="{width_emu}"><a:solidFill><a:srgbClr val="{color}"/></a:solidFill></a:ln>'
    pattern = re.compile(r'(<spPr><a:prstGeom prst="rect"\s*/>)(</spPr>)')

    tmp_path = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                text = pattern.sub(lambda m: m.group(1) + ln_xml + m.group(2), text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp_path, path)


def resolve_insight_output_path(
    story_name: str,
    filename: str | None = None,
    *,
    override_abs_dir: str | Path | None = None,
    repo_root: Path | None = None,
) -> Path:
    """
    Resolve the save path for an insight workbook.

    Default: <repo_root>/story/<story_name>/output/InsightResult_<story_name>.xlsx
    (dir created if missing) — repo-local, alongside the chart PNGs it references.

    Pass override_abs_dir (e.g. a OneDrive/CDDA business path) to save there
    instead, for handing the deliverable off to stakeholders. No business path
    is hardcoded here — see libs/.claude/skills/insight-draft/ for that variant.
    """
    if filename is None:
        filename = f"InsightResult_{story_name}.xlsx"

    if override_abs_dir is not None:
        out_dir = Path(override_abs_dir)
    else:
        if repo_root is None:
            repo_root = Path(__file__).resolve().parent.parent.parent
        out_dir = repo_root / "story" / story_name / "output"

    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / filename


def save(wb: Workbook, path: str | Path) -> None:
    """
    Save the workbook to `path`. If a file already exists at that path, it
    is backed up first (renamed in place with a timestamp suffix) before the
    new version is written, so re-running the builder after edits never
    silently overwrites the previous draft. Every pasted picture then gets a
    native Picture Border patched into its drawing XML (see
    `_add_picture_borders`), since openpyxl can't set that at insert time.
    """
    path = str(path)
    if os.path.exists(path):
        stem, ext = os.path.splitext(path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{stem}_backup_{timestamp}{ext}"
        shutil.move(path, backup_path)
    wb.save(path)
    _add_picture_borders(path)
